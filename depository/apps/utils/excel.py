import os
from uuid import uuid4

from django.conf import settings
from openpyxl import load_workbook, Workbook


class ExcelUtil(object):
    def import_file(self, path):
        """
        import data from excel file
        :param path: file path
        :return: list of dictionary
        """
        data = []
        wb = load_workbook(path, read_only=True)
        sheet = wb.active

        header = next(sheet.rows)

        columns = []
        for col in header:
            columns.append(col.value)

        for row in sheet.rows:
            data.append({name: cell.value for name, cell in zip(columns, row)})

        return data

    def export_file(self, header, data):
        """
        export given data to the path
        :param header: list of string that includes name of columns
        :param data: list of dictionary
        :return: a temporary file
        """
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Depository'

        # write header
        for idx, col_name in enumerate(header, start=1):
            _ = sheet.cell(column=idx, row=1, value=col_name)

        # write data
        for row_idx, row in enumerate(data, start=2):
            for col_idx, col_name in enumerate(header, start=1):
                cell_value = row.get(col_name, '')
                _ = sheet.cell(column=col_idx, row=row_idx, value=str(cell_value))

        file_name = f'{uuid4().hex}.csv'
        path = os.path.join(settings.EXPORT_ROOT, file_name)
        wb.save(path)

        return file_name
